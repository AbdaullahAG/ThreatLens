"""
Excel Reporter — generates a professional, color-coded Excel workbook.
"""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from src.models import EnrichmentResult, IOCType


# ── Color palette ───────────────────────────────────────────────────────────
C_RED        = "FFDC143C"
C_ORANGE     = "FFFF8C00"
C_YELLOW     = "FFFFD700"
C_GREEN      = "FF228B22"
C_BLUE       = "FF1E90FF"
C_DARK       = "FF1A1A2E"
C_MID        = "FF16213E"
C_LIGHT_GRAY = "FFF5F5F5"
C_WHITE      = "FFFFFFFF"
C_HEADER_BG  = "FF0F3460"
C_HEADER_FG  = "FFFFFFFF"

VERDICT_COLORS = {
    "Malicious": C_RED,
    "Suspicious": C_ORANGE,
    "Clean":      C_GREEN,
    "Unknown":    "FF808080",
    "Critical":   C_RED,
    "High":       C_ORANGE,
    "Medium":     C_YELLOW,
    "Low":        C_GREEN,
}

THIN_BORDER = Border(
    left=Side(style="thin", color="FFD0D0D0"),
    right=Side(style="thin", color="FFD0D0D0"),
    top=Side(style="thin", color="FFD0D0D0"),
    bottom=Side(style="thin", color="FFD0D0D0"),
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _bold_font(size: int = 11, color: str = "FF000000") -> Font:
    return Font(bold=True, size=size, color=color)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


class ExcelReporter:

    def generate(
        self,
        results: List[EnrichmentResult],
        output_dir: str = "output",
    ) -> str:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Summary Dashboard
        self._build_summary(wb, results)

        # Per-type sheets
        ip_results      = [r for r in results if r.ioc.ioc_type == IOCType.IP]
        domain_results  = [r for r in results if r.ioc.ioc_type in (IOCType.DOMAIN, IOCType.URL)]
        hash_results    = [r for r in results if r.ioc.ioc_type == IOCType.HASH]
        cve_results     = [r for r in results if r.ioc.ioc_type == IOCType.CVE]

        if ip_results:
            self._build_ip_sheet(wb, ip_results)
        if domain_results:
            self._build_domain_sheet(wb, domain_results)
        if hash_results:
            self._build_hash_sheet(wb, hash_results)
        if cve_results:
            self._build_cve_sheet(wb, cve_results)

        # Save
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = str(Path(output_dir) / f"ThreatLens_Report_{ts}.xlsx")
        wb.save(filepath)
        return filepath

    # ── Summary Sheet ────────────────────────────────────────────────────────

    def _build_summary(self, wb: Workbook, results: List[EnrichmentResult]):
        ws = wb.create_sheet("📊 Summary")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 22

        # Title
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = "🔍 ThreatLens — Threat Intelligence Report"
        title_cell.font = Font(bold=True, size=16, color=C_WHITE)
        title_cell.fill = _fill(C_HEADER_BG)
        title_cell.alignment = _center()
        ws.row_dimensions[1].height = 36

        # Timestamp
        ws.merge_cells("A2:E2")
        ts_cell = ws["A2"]
        ts_cell.value = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ts_cell.font = Font(italic=True, size=10, color="FF888888")
        ts_cell.alignment = _center()

        ws.row_dimensions[3].height = 10  # spacer

        # Stats header
        headers = ["IOC Type", "Total", "Malicious", "Suspicious", "Clean"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = _bold_font(11, C_WHITE)
            cell.fill = _fill(C_MID)
            cell.alignment = _center()
            cell.border = THIN_BORDER

        # Stats rows
        ioc_types = [
            (IOCType.IP, "🌐 IP Addresses"),
            (IOCType.DOMAIN, "🔗 Domains/URLs"),
            (IOCType.URL, None),
            (IOCType.HASH, "🔑 File Hashes"),
            (IOCType.CVE, "⚠️  CVEs"),
        ]

        row = 5
        for ioc_type, label in ioc_types:
            if label is None:
                continue
            type_results = [
                r for r in results
                if r.ioc.ioc_type == ioc_type
                or (ioc_type == IOCType.DOMAIN and r.ioc.ioc_type == IOCType.URL)
            ]
            if not type_results:
                continue

            total     = len(type_results)
            malicious = sum(1 for r in type_results if r.verdict in ("Malicious", "Critical"))
            suspicious = sum(1 for r in type_results if r.verdict in ("Suspicious", "High"))
            clean     = total - malicious - suspicious

            row_data = [label, total, malicious, suspicious, clean]
            fills = [None, None, C_RED if malicious else None, C_ORANGE if suspicious else None, C_GREEN if clean else None]

            for col, (val, fill) in enumerate(zip(row_data, fills), 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = _center()
                cell.border = THIN_BORDER
                if fill:
                    cell.fill = _fill(fill)
                    cell.font = Font(bold=True, color=C_WHITE)
                elif col == 1:
                    cell.font = _bold_font()
                    cell.fill = _fill(C_LIGHT_GRAY)
            row += 1

        # All IOCs table
        row += 2
        ws.cell(row=row, column=1, value="All Investigated IOCs").font = _bold_font(12)
        row += 1

        all_headers = ["IOC Value", "Type", "Verdict", "Country", "Key Finding"]
        for col, h in enumerate(all_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = _bold_font(10, C_WHITE)
            cell.fill = _fill(C_HEADER_BG)
            cell.alignment = _center()
            cell.border = THIN_BORDER

        row += 1
        for r in results:
            key_finding = _key_finding(r)
            row_vals = [r.ioc.value, r.ioc.ioc_type.value, r.verdict, r.country or "—", key_finding]
            for col, val in enumerate(row_vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = _left()
                cell.border = THIN_BORDER
                if col == 3:
                    color = VERDICT_COLORS.get(r.verdict, "FF808080")
                    cell.fill = _fill(color)
                    cell.font = Font(bold=True, color=C_WHITE)
            row += 1

    # ── IP Sheet ─────────────────────────────────────────────────────────────

    def _build_ip_sheet(self, wb: Workbook, results: List[EnrichmentResult]):
        ws = wb.create_sheet("🌐 IP Addresses")
        ws.sheet_view.showGridLines = False

        headers = [
            "IP Address", "Verdict", "Abuse Score", "Country", "ISP/Org",
            "Usage Type", "Total Reports", "Last Reported",
            "Open Ports", "Shodan Vulns", "Hostnames",
            "VT Malicious", "OTX Pulses", "ASN", "Errors",
        ]
        col_widths = [18, 12, 14, 12, 24, 18, 15, 20, 22, 22, 28, 14, 12, 16, 30]
        self._write_header_row(ws, headers, col_widths)

        for row_idx, r in enumerate(results, 2):
            s = r.sources
            row_vals = [
                r.ioc.value,
                r.verdict,
                r.abuse_score,
                r.country or "—",
                r.organization or r.isp or "—",
                r.usage_type or "—",
                r.total_reports,
                r.last_reported or "—",
                ", ".join(str(p) for p in r.open_ports) or "—",
                ", ".join(r.shodan_vulns) or "—",
                ", ".join(r.hostnames[:5]) or "—",
                s.get("virustotal", {}).get("malicious", "—"),
                s.get("otx", {}).get("pulse_count", "—"),
                r.asn or "—",
                "; ".join(f"{k}: {v}" for k, v in r.errors.items()) or "—",
            ]
            self._write_data_row(ws, row_idx, row_vals, verdict=r.verdict, verdict_col=2)

    # ── Domain/URL Sheet ──────────────────────────────────────────────────────

    def _build_domain_sheet(self, wb: Workbook, results: List[EnrichmentResult]):
        ws = wb.create_sheet("🔗 Domains & URLs")
        ws.sheet_view.showGridLines = False

        headers = [
            "Domain / URL", "Type", "Verdict", "VT Malicious",
            "VT Suspicious", "VT Harmless", "Categories",
            "URLScan Verdict", "Tags", "Country", "Errors",
        ]
        col_widths = [40, 10, 12, 14, 14, 12, 28, 16, 22, 12, 30]
        self._write_header_row(ws, headers, col_widths)

        for row_idx, r in enumerate(results, 2):
            row_vals = [
                r.ioc.value,
                r.ioc.ioc_type.value,
                r.verdict,
                r.malicious_votes,
                r.suspicious_votes,
                r.harmless_votes,
                ", ".join(r.categories) or "—",
                r.urlscan_verdict or "—",
                ", ".join(r.tags[:5]) or "—",
                r.country or "—",
                "; ".join(f"{k}: {v}" for k, v in r.errors.items()) or "—",
            ]
            self._write_data_row(ws, row_idx, row_vals, verdict=r.verdict, verdict_col=3)

    # ── Hash Sheet ────────────────────────────────────────────────────────────

    def _build_hash_sheet(self, wb: Workbook, results: List[EnrichmentResult]):
        ws = wb.create_sheet("🔑 File Hashes")
        ws.sheet_view.showGridLines = False

        headers = [
            "Hash", "Verdict", "File Name", "File Type",
            "File Size (bytes)", "Detections", "Total Scanners",
            "Detection Rate", "Tags", "OTX Pulses", "Errors",
        ]
        col_widths = [68, 12, 28, 18, 18, 14, 16, 16, 30, 12, 30]
        self._write_header_row(ws, headers, col_widths)

        for row_idx, r in enumerate(results, 2):
            det_rate = (
                f"{r.positives}/{r.total_scanners} "
                f"({r.positives*100//r.total_scanners}%)"
                if r.total_scanners else "—"
            )
            row_vals = [
                r.ioc.value,
                r.verdict,
                r.file_name or "—",
                r.file_type or "—",
                r.file_size or "—",
                r.positives,
                r.total_scanners,
                det_rate,
                ", ".join(r.tags[:5]) or "—",
                r.sources.get("otx", {}).get("pulse_count", "—"),
                "; ".join(f"{k}: {v}" for k, v in r.errors.items()) or "—",
            ]
            self._write_data_row(ws, row_idx, row_vals, verdict=r.verdict, verdict_col=2)

    # ── CVE Sheet ─────────────────────────────────────────────────────────────

    def _build_cve_sheet(self, wb: Workbook, results: List[EnrichmentResult]):
        ws = wb.create_sheet("⚠️ CVEs")
        ws.sheet_view.showGridLines = False

        headers = [
            "CVE ID", "Severity", "CVSS Score", "CVSS Version",
            "Published Date", "Description",
            "Affected Products (sample)", "References (sample)",
        ]
        col_widths = [18, 12, 13, 14, 16, 60, 40, 60]
        self._write_header_row(ws, headers, col_widths)

        for row_idx, r in enumerate(results, 2):
            row_vals = [
                r.ioc.value,
                r.severity or "Unknown",
                r.cvss_score or "—",
                r.cvss_version or "—",
                r.published_date or "—",
                r.cve_description or "—",
                "; ".join(r.affected_products[:5]) or "—",
                "\n".join(r.references[:3]) or "—",
            ]
            self._write_data_row(ws, row_idx, row_vals, verdict=r.verdict, verdict_col=2)
            ws.row_dimensions[row_idx].height = 60

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _write_header_row(self, ws, headers: list, col_widths: list):
        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = _bold_font(10, C_WHITE)
            cell.fill = _fill(C_HEADER_BG)
            cell.alignment = _center()
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

    def _write_data_row(self, ws, row_idx: int, values: list, verdict: str, verdict_col: int):
        alt = row_idx % 2 == 0
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = _left()
            if col == verdict_col:
                color = VERDICT_COLORS.get(verdict, "FF808080")
                cell.fill = _fill(color)
                cell.font = Font(bold=True, color=C_WHITE)
                cell.alignment = _center()
            elif alt:
                cell.fill = _fill(C_LIGHT_GRAY)
        ws.row_dimensions[row_idx].height = 18


def _key_finding(r: EnrichmentResult) -> str:
    if r.ioc.ioc_type == IOCType.IP:
        parts = []
        if r.abuse_score:
            parts.append(f"Abuse: {r.abuse_score}%")
        if r.open_ports:
            parts.append(f"Ports: {', '.join(str(p) for p in r.open_ports[:4])}")
        if r.shodan_vulns:
            parts.append(f"Vulns: {', '.join(r.shodan_vulns[:2])}")
        return "; ".join(parts) or "No notable findings"
    elif r.ioc.ioc_type in (IOCType.DOMAIN, IOCType.URL):
        return f"VT Malicious: {r.malicious_votes} | URLScan: {r.urlscan_verdict or 'N/A'}"
    elif r.ioc.ioc_type == IOCType.HASH:
        return f"Detections: {r.positives}/{r.total_scanners}" if r.total_scanners else "No scan data"
    elif r.ioc.ioc_type == IOCType.CVE:
        return f"CVSS: {r.cvss_score} ({r.severity})" if r.cvss_score else "No CVSS data"
    return "—"
